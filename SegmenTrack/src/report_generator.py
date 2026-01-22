import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

class ReportGenerator:
    def generate_report(self, analysis_results, plot_paths, output_path):
        df = pd.DataFrame.from_dict(analysis_results, orient="index")

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Summary")

        wb = load_workbook(output_path)
        ws = wb.create_sheet("Charts")

        for img_path in plot_paths:
            img = Image(img_path)
            ws.add_image(img, "A1")

        wb.save(output_path)
